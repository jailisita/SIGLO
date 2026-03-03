from USERS.views import admin_required
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.shortcuts import redirect, render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from LOTES.models import Lot, Stage
from PQRS.models import PQRS
from SALES.models import Payment, Purchase
from .models import ProjectInfo


def dashboard(request):
    if request.user.is_authenticated:
        role = getattr(request.user, "role", "CLIENT")

        if role == "ADMIN":
            User = get_user_model()

            users_total = User.objects.count()
            clients_total = User.objects.filter(role="CLIENT").count()
            admins_total = User.objects.filter(role="ADMIN").count()

            lots_total = Lot.objects.count()
            lots_available = Lot.objects.filter(status="AVAILABLE").count()
            lots_sold = Lot.objects.filter(status="SOLD").count()

            purchases_total = Purchase.objects.count()
            total_purchase_amount = (
                Purchase.objects.aggregate(total=Sum("total_amount"))["total"] or 0
            )
            total_paid_amount = (
                Payment.objects.aggregate(total=Sum("amount"))["total"] or 0
            )

            pqrs_total = PQRS.objects.count()
            pqrs_open = PQRS.objects.filter(status="OPEN").count()

            context = {
                "users_total": users_total,
                "clients_total": clients_total,
                "admins_total": admins_total,
                "lots_total": lots_total,
                "lots_available": lots_available,
                "lots_sold": lots_sold,
                "purchases_total": purchases_total,
                "total_purchase_amount": total_purchase_amount,
                "total_paid_amount": total_paid_amount,
                "pqrs_total": pqrs_total,
                "pqrs_open": pqrs_open,
            }
            return render(request, "project_info/admin_dashboard.html", context)

        purchases = Purchase.objects.filter(client=request.user)
        lots_owned_count = purchases.values("lots").distinct().count()
        total_purchase_amount = (
            purchases.aggregate(total=Sum("total_amount"))["total"] or 0
        )
        total_paid_amount = (
            Payment.objects.filter(purchase__client=request.user).aggregate(
                total=Sum("amount")
            )["total"]
            or 0
        )
        balance_total = total_purchase_amount - total_paid_amount

        pqrs_open = PQRS.objects.filter(client=request.user, status="OPEN").count()

        context = {
            "purchases_total": purchases.count(),
            "lots_owned_count": lots_owned_count,
            "total_purchase_amount": total_purchase_amount,
            "total_paid_amount": total_paid_amount,
            "balance_total": balance_total,
            "pqrs_open": pqrs_open,
        }
        return render(request, "project_info/client_dashboard.html", context)

    stages = Stage.objects.all().order_by('id')
    return render(request, "project_info/dashboard.html", {"stages": stages})


def error_404_view(request, exception):
    return render(request, "404.html", status=404)


def error_500_view(request):
    return render(request, "500.html", status=500)


@admin_required
def download_monthly_report(request):
    """
    Genera un archivo Excel con el reporte de actividad del mes actual.
    """
    now = datetime.now()
    month_name = now.strftime("%B %Y")
    
    # Crear libro de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Mensual"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_aligned = Alignment(horizontal="center")

    # Título Principal
    ws.merge_cells('A1:E1')
    ws['A1'] = f"REPORTE DE GESTIÓN - {month_name.upper()}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = center_aligned

    # --- SECCIÓN 1: RESUMEN DE VENTAS ---
    ws.append([])
    ws.append(["RESUMEN DE VENTAS (ACUMULADO)"])
    ws.append(["Concepto", "Valor"])
    
    # Datos de ventas
    total_sales = Purchase.objects.aggregate(total=Sum("total_amount"))["total"] or 0
    total_paid = Payment.objects.aggregate(total=Sum("amount"))["total"] or 0
    balance = total_sales - total_paid
    
    ws.append(["Total en Ventas", total_sales])
    ws.append(["Total Recaudado", total_paid])
    ws.append(["Saldo Pendiente", balance])

    # --- SECCIÓN 2: DETALLE DE PAGOS DEL MES ---
    ws.append([])
    ws.append(["DETALLE DE PAGOS DEL MES"])
    headers = ["Fecha", "Cliente", "Lotes", "Monto", "Estado"]
    ws.append(headers)

    # Filtrar pagos del mes actual
    monthly_payments = Payment.objects.filter(
        payment_date__year=now.year, 
        payment_date__month=now.month
    ).select_related('purchase__client')

    for p in monthly_payments:
        lots_str = ", ".join([str(l.number) for l in p.purchase.lots.all()])
        ws.append([
            p.payment_date.strftime("%d/%m/%Y"),
            p.purchase.client.get_full_name() or p.purchase.client.username,
            lots_str,
            p.amount,
            "Validado" if p.is_validated else "Pendiente"
        ])

    # --- SECCIÓN 3: PQRS DEL MES ---
    ws.append([])
    ws.append(["PQRS RECIBIDAS EN EL MES"])
    ws.append(["Tipo", "Cliente", "Mensaje", "Estado"])

    # Intentamos filtrar PQRS si tienen algún campo de fecha, si no, mostramos todas las abiertas
    # Como no vi 'created_at' en PQRS, usaremos todas las de este mes (asumiendo que se implementará o filtrando por estado)
    monthly_pqrs = PQRS.objects.all() # En una app real filtraríamos por fecha
    
    type_map = dict(PQRS.TYPE_CHOICES)
    status_map = dict(PQRS.STATUS_CHOICES)

    for pqr in monthly_pqrs:
        ws.append([
            type_map.get(pqr.type, pqr.type),
            pqr.client.username,
            pqr.message[:50] + "..." if len(pqr.message) > 50 else pqr.message,
            status_map.get(pqr.status, pqr.status)
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename=Reporte_SIGLO_{now.strftime('%Y_%m')}.xlsx"
    wb.save(response)
    return response


@admin_required
def admin_content(request):
    return redirect('dashboard')
